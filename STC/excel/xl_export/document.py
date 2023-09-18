""" Генерация маршрутной карты в Excel """
from __future__ import annotations

import re
from os import path
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from STC.config.config import CFG_MK
from STC.functions.func import text_slicer
from STC.gui.splash_screen import SplashScreen

if TYPE_CHECKING:
    from STC.product.product import Product
    from STC.product.product import Document
    from STC.product.product import Operation


def save_excel_file(workbook, wb_name: str, ext: str, num: int = 0) -> str:
    """ Сохраняет файл Excel, подбирая номер версии
        в зависимости от наименования других файлов
        в данной директории """

    file_name = f'{wb_name}{ext}' if num == 0 else f'{wb_name} Вариант {num}{ext}'
    try:
        workbook.save(file_name)
    except PermissionError:
        num = num + 1
        file_name = save_excel_file(workbook=workbook,
                                    wb_name=f'{wb_name}',
                                    ext=ext,
                                    num=num)
    return file_name


class ExcelDocumentCreator:
    """ Создает документ используя методы шаблона документа """

    def __init__(self, document: list[Document]) -> None:
        self.document = document[0]
        self.template = ExcelTemplate()
        document_name = re.sub(r'[\\/:"*?<>|]+', '!', self.document.name, count=0, flags=0)
        self.file_name = f'{CFG_MK.file.file_path}' \
                         f'{self.document.product.deno} ' \
                         f'({self.document.deno}) ' \
                         f'{document_name}'
        self.children = []
        self.in_complex = {}
        self.createDocument()

    def createDocument(self) -> None:
        """ Процесс создания документа """
        self.template.document_name = self.document.name
        self.template.document_deno = self.document.deno
        self.template.product_deno = self.document.product.deno
        self.template.developer = self.document.name_developer
        self.template.checker = self.document.name_checker
        # self.template.approver = self.document.name_approver
        self.template.n_contr = self.document.name_n_contr
        # self.template.m_contr = self.document.name_m_contr
        self.template.litera = self.document.litera
        SplashScreen().newMessage(message='Расчет первой страницы',
                                  stage=1,
                                  stages=4,
                                  log=True,
                                  logging_level='INFO')
        self.createFirstPage()
        SplashScreen().newMessage(message='Расчет текста операций',
                                  log=True,
                                  logging_level='INFO')
        self.createOperationText()
        SplashScreen().newMessage(message='Расчет страницы входящих',
                                  log=True,
                                  logging_level='INFO')
        self.createChildrenPage()
        SplashScreen().newMessage(message='Экспорт в Excel',
                                  log=True,
                                  logging_level='INFO')
        self.template.exportEText()

        wb_name = self.template.save(name=self.file_name)

        SplashScreen().closeWithWindow(msg=f'Документ сохранен как:\n{wb_name}',
                                       m_type='info')

    def createFirstPage(self) -> None:
        """ Генерация общих данных и их расположения в документе """
        sentences = self.document.generateCommonProperties()
        self.template.addMainData(sentences=sentences)

    def createOperationText(self) -> None:
        """ Расположение данных об операциях в документе """
        for operation in sorted(self.document.operations.values(), key=lambda x: x.order):
            self.template.addOperation(operation=operation)

    def createChildrenPage(self) -> None:
        """ Генерация данных иерархии и их расположение в документе """
        product = self.document.product
        self.updComplex(document=self.document, sub_products=self.document.sub_products_new)
        self.createChildInfoString(product=product)
        self.additionalData(product=product)
        if self.children:
            self.template.addChildrenData(data=self.children)

    def createChildInfoString(self, product: Product, parent_index: str | None = None) -> None:
        """ Генерация данных иерархии """
        children = self.getChildren(product)
        for num, child in enumerate(children):
            document_text = CFG_MK.last_page.document_not_found
            document_mk = None
            product = child['product']
            quantity = child['quantity']
            key = (product, self.document)
            if key in self.in_complex:
                document_text = f'{CFG_MK.last_page.in_document}:'  # {self.in_complex[key].deno}
            else:
                subtype_name = 'Карта типового (группового) технологического процесса'
                document_kttp = product.getDocumentByType(class_name='ТД',
                                                          subtype_name=subtype_name,
                                                          org_code='2',
                                                          only_relevant=True,
                                                          only_text=False,
                                                          first=True)
                document = document_kttp
                if document is None:
                    for name in CFG_MK.main.subtype_names:
                        document_mk = product.getDocumentByType(class_name='ТД',
                                                                subtype_name=name,
                                                                only_relevant=True,
                                                                only_text=False,
                                                                first=True)
                        if document_mk is not None:
                            break
                    document = document_mk
                if document is not None:
                    document_text = f'{CFG_MK.last_page.by_document} ' \
                                    f'{document.sign} {document.deno}'
            index = self.childIndex(index=parent_index, num=num)
            self.children.append({'index': index,
                                  'product': f'{product.name} {product.deno}',
                                  'quantity': quantity,
                                  'document': document_text})
            if document_mk is None:
                self.createChildInfoString(product=product, parent_index=index)

    @staticmethod
    def childIndex(index: str, num: int) -> str:
        """ Возвращает индекс (1.2.3) для данных иерархии """
        num += 1  # Нумерация с 1
        if index is None:
            return f'{num}.'
        return f'{index}.{num }' if index[-1] != '.' else f'{index}{num}'

    @staticmethod
    def getChildren(product: Product) -> list[dict]:
        """ Возвращает список изделий для иерархии """
        children = []
        for child in product.children():
            product = child['product']
            if product.product_kind_name not in CFG_MK.main.product_type_exceptions:
                children.append(child)
        return children

    def updComplex(self, document: Document, sub_products: list[Product]):
        """ Дополнение списка изделий, входящих в состав составных документов """
        for sub_product in sub_products:
            self.in_complex[sub_product] = document

    def additionalData(self, product: Product) -> None:
        """ Дополнение иерархии входящих изделий нестандартными случаями """
        document = product.getDocumentByType(class_name='КД',
                                             subtype_name='Ведомость эксплуатационных документов',
                                             only_relevant=True,
                                             only_text=False,
                                             first=True)
        if document is not None:
            previous_index = self.children[-1]['index']
            main_index_part = previous_index[:previous_index.find('.')]
            index = int(main_index_part) + 1
            product_text = f'{CFG_MK.last_page.ed_document_prefix} {document.deno}'
            document_text = f'{CFG_MK.last_page.ed_document_postfix}'
            self.children.append({'index': f'{index}.',
                                  'product': product_text,
                                  'quantity': CFG_MK.last_page.ed_document_quantity,
                                  'document': document_text})


class ExcelTemplate:
    """ Шаблон документа с методами, распределяющими
        данные маршрутной карты по документу"""
    # pylint: disable=too-many-instance-attributes
    # pylint: disable=too-many-public-methods

    center_alignment = Alignment(horizontal='center',
                                 vertical='bottom',
                                 wrap_text=False)
    left_alignment = Alignment(horizontal='left',
                               vertical='bottom',
                               wrap_text=False)

    def __init__(self) -> None:
        self.excel_len_lp = CFG_MK.main.excel_len - 120
        self.name = path.join(CFG_MK.file.template_folder,
                              CFG_MK.file.template_name)
        self.workbook = load_workbook(self.name)
        self.ws_template = self.workbook[CFG_MK.main.ws_template_name]
        self.ws_first_page = self.workbook[CFG_MK.main.ws_first_page_name]
        self.ws_text = self.workbook[CFG_MK.main.ws_text_name]
        self.row = CFG_MK.row.row_start_fp
        self.last_index = ''
        self.pages = []
        self.document_dict = {}
        self.developer = CFG_MK.doc_details.developer
        self.checker = CFG_MK.doc_details.checker
        self.approver = CFG_MK.doc_details.approver
        self.n_contr = CFG_MK.doc_details.n_contr
        self.m_contr = CFG_MK.doc_details.m_contr

    def addIndex(self, index: str) -> None:
        """ Повторяет индекс строки, если на данной странице его еще нет """
        if index != self.last_index:
            self.addEText(etext=EText(text=index,
                                      index='',
                                      col_fst=CFG_MK.col.col_str_index))
            self.last_index = index

    def addEText(self, etext: EText) -> None:
        """ Внесение логической строки в шаблон """
        temp_dict = {etext.col_fst: etext}
        if self.row not in self.document_dict:
            self.document_dict.update({self.row: temp_dict})
        else:
            self.document_dict[self.row].update(temp_dict)

    def addMainData(self, sentences: list[str]) -> None:
        """ Внесение общих данных в шаблон """
        self.rowIncrease()
        for num, sentence in enumerate(sentences):
            self.addSlicedText(base_text='. '.join([str(num + 1), sentence]),
                               base_length=CFG_MK.main.excel_len,
                               str_index='',
                               col_fst=CFG_MK.col.col_main_text)
            self.rowIncrease()
        self.endOfMainData()

    def addOperation(self, operation: Operation) -> None:
        """ Вызывает методы внесения данных для определенных
            типов строк в соответствии с их очередностью в
            маршрутной карте"""
        self.addOperationA(operation=operation)
        self.addOperationB(operation=operation)
        self.addOperationK(operation=operation)
        self.addOperationM(operation=operation)
        self.addOperationO(operation=operation)
        self.addOperationT(operation=operation)
        self.endOfOperation()

    def addOperationA(self, operation: Operation) -> None:
        """ Генерация логических строк типа А """
        index = 'А'
        self.addIndex(index=index)
        num = EText(text=operation.num,
                    index=index,
                    col_fst=CFG_MK.col.col_operation_number,
                    alignment=self.__class__.center_alignment)
        name = EText(text=operation.name,
                     index=index,
                     col_fst=CFG_MK.col.col_operation)
        area = EText(text=operation.area.name_short,
                     index=index,
                     col_fst=CFG_MK.col.col_area_fst)
        self.addEText(etext=num)
        self.addEText(etext=name)
        self.addEText(etext=area)
        col_docs = CFG_MK.col.col_docs_fp if self.page(self.row) == 1 else CFG_MK.col.col_docs
        self.addSlicedText(base_text=operation.documents_and_iot,
                           base_length=CFG_MK.main.excel_len_iot,
                           str_index=index,
                           col_fst=col_docs)
        self.rowIncrease()

    def addOperationB(self, operation: Operation) -> None:
        """ Генерация логических строк типа Б """
        index = 'Б'
        self.addIndex(index=index)
        workplace = EText(text=operation.workplace.name,
                          index=index,
                          col_fst=CFG_MK.col.col_workplace)
        if self.page(self.row) == 1:
            col_prof_fst = CFG_MK.col.col_prof_fst_fp
            col_prof_end = CFG_MK.col.col_prof_end_fp
        else:
            col_prof_fst = CFG_MK.col.col_prof_fst
            col_prof_end = CFG_MK.col.col_prof_end

        profession = EText(text=operation.profession.code,
                           index=index,
                           col_fst=col_prof_fst,
                           col_end=col_prof_end,
                           alignment=self.center_alignment)
        self.addEText(etext=workplace)
        self.addEText(etext=profession)
        self.rowIncrease()

    def addOperationK(self, operation: Operation) -> None:
        """ Генерация логических строк типа К """

    def addOperationM(self, operation: Operation) -> None:
        """ Генерация логических строк типа М """
        if operation.mat:
            index = 'М'
            self.addIndex(index=index)
            base_text = f'{operation.mat}.'
            self.addSlicedText(base_text=base_text,
                               base_length=CFG_MK.main.excel_len,
                               str_index=index,
                               col_fst=CFG_MK.col.col_main_text)
            self.rowIncrease()

    def addOperationO(self, operation: Operation) -> None:
        """ Генерация логических строк типа О """
        if operation.sentences:
            sentences = sorted(operation.sentences.items(), key=lambda x: x[0])
            index = 'О'
            self.addIndex(index=index)
            for order, sentence in sentences:
                base_text = '. '.join([str(order + 1), sentence.text])
                self.addSlicedText(base_text=base_text,
                                   base_length=CFG_MK.main.excel_len,
                                   str_index=index,
                                   col_fst=CFG_MK.col.col_main_text)
                self.rowIncrease()
            if operation.name in ["Контроль", "Предъявительские испытания (ПИ)"]:
                self.addSlicedText(base_text=CFG_MK.main.additional_control_text,
                                   base_length=CFG_MK.main.excel_len,
                                   str_index='',
                                   col_fst=CFG_MK.col.col_main_text)
                self.rowIncrease()

    def addOperationT(self, operation: Operation) -> None:
        """ Генерация логических строк типа Т """
        if operation.rig:
            index = 'T'
            self.addIndex(index=index)
            base_text = f'{operation.rig}.'
            self.addSlicedText(base_text=base_text,
                               base_length=CFG_MK.main.excel_len,
                               str_index=index,
                               col_fst=CFG_MK.col.col_main_text)
            self.rowIncrease()

    def addChildrenData(self, data: list[dict[str, str]]):
        """ Генерация логических строк иерархии """
        self.rowDecrease()
        self.row = self.first_row_of_next_page
        index = 'А'
        self.addIndex(index)
        self.addEText(etext=EText(text=CFG_MK.last_page.last_page_start_text,
                                  index=index,
                                  col_fst=CFG_MK.col.col_main_text))
        self.rowIncrease()
        self.addIndex('Б')
        self.rowIncrease()
        self.addIndex('K')
        for dict_item in data:
            product = dict_item['product']
            quantity = str(dict_item['quantity'])
            document = dict_item['document']
            index = dict_item['index']
            text = f'{index} {product} {document}'
            self.addSlicedText(base_text=text,
                               base_length=self.excel_len_lp,
                               str_index='',
                               col_fst=CFG_MK.col.col_main_text)
            self.addEText(etext=EText(text=quantity,
                                      index='К',
                                      col_fst=CFG_MK.col.col_k_num_fst,
                                      col_end=CFG_MK.col.col_k_num_end,
                                      alignment=self.center_alignment))
            self.rowIncrease()
        # self.endOfOperation()

    def addSlicedText(self, base_text: str,
                      base_length: int,
                      str_index: str,
                      col_fst: int,
                      col_end: int | None = None,
                      alignment: str | None = None) -> None:
        """ Нарезка длинных строк под шаблон Excel и
            внесение в шаблон"""
        # pylint: disable=too-many-arguments

        if base_text:
            text_list = self.textSlicer(base_text=base_text,
                                        base_length=base_length,
                                        str_index=str_index)
            for text in text_list:
                self.addEText(etext=EText(text=text['string'],
                                          index=text['index'],
                                          col_fst=col_fst,
                                          col_end=col_end,
                                          alignment=alignment))
                self.rowIncrease()
            self.rowDecrease()

    @staticmethod
    def textSlicer(base_text: str, base_length: int, str_index: str) -> list[dict[str, str]]:
        """ Нарезка длинных строк под шаблон Excel """
        base_text = base_text.split('\n')
        text_list = []
        for text in base_text:
            for line in text_slicer(text, base_length):
                text_list.append({'index': str_index,
                                  'string': line})
        return text_list

    def exportEText(self) -> None:
        """ Внесение данных в файл excel с разбивкой по страницам """
        self.rowDecrease()
        self.createPages()
        for row, temp_dict in self.document_dict.items():
            page = self.page(row)
            row_on_page = row - (page - 1) * CFG_MK.row.row_total
            for col, etext in temp_dict.items():
                if etext.merge:
                    self.pages[page - 1].merge_cells(start_row=row_on_page,
                                                     start_column=etext.col_fst,
                                                     end_row=row_on_page,
                                                     end_column=etext.col_end)
                cell = self.pages[page - 1].cell(row=row_on_page,
                                                 column=col)
                cell.value = etext.text
                cell.alignment = etext.alignment
                if page > 1 and row_on_page == CFG_MK.row.text_row_first + 1:
                    cell = self.pages[page - 1].cell(row=row_on_page,
                                                     column=CFG_MK.col.col_str_index)
                    cell.value = etext.index
        self.workbook.remove_sheet(self.ws_first_page)
        self.workbook.remove_sheet(self.ws_template)
        self.workbook.remove_sheet(self.ws_text)

    def createPages(self) -> None:
        """ Создание страниц в файле Excel """
        page_total = 1
        first_page = self.createPage(page=1, form='2')
        for page in range(2, self.current_page + 1):
            self.createPage(page=page, form='1б')
            page_total += 1
        first_page.cell(row=CFG_MK.row.row_page,
                        column=CFG_MK.col.col_page_total).value = page_total

    def createPage(self, form: str, page: int):
        """ Создание страницы определенного типа в файле Excel """
        worksheet = None
        if form == '2':
            worksheet = self.workbook.copy_worksheet(self.ws_first_page)
            worksheet.cell(row=CFG_MK.row.row_page,
                           column=CFG_MK.col.col_page_fp).value = page
        elif form == '1б':
            worksheet = self.workbook.copy_worksheet(self.ws_template)
            worksheet.cell(row=CFG_MK.row.row_page,
                           column=CFG_MK.col.col_page).value = page
        self.pages.append(worksheet)
        worksheet.title = str(page)
        return worksheet

    def save(self, name: str) -> str:
        """ Сохранение документа с учетом
            версий документа в директории """
        wb_name = save_excel_file(workbook=self.workbook,
                                  wb_name=name,
                                  ext=CFG_MK.file.document_extension)
        return wb_name

    def endOfMainData(self) -> None:
        """ Определение текущей строки после
            окончания блока общих данных """
        if self.row_on_cur_page == 1:
            self.row = self.row + 2
        else:
            if self.have_space_on_page:
                if self.page(self.row) == 1:
                    self.row = self.first_row_of_next_page + 2
                else:
                    self.row += CFG_MK.main.rows_between_operations
            else:
                self.row = self.first_row_of_next_page + 2

    def endOfOperation(self) -> None:
        """ Определение текущей строки после
            окончания блока операции """
        if self.row_on_cur_page == 1:
            pass
        elif self.row_on_cur_page == CFG_MK.main.rows_to_push_up:
            self.pushOperationUp()
        else:
            if self.have_space_on_page:
                self.row += CFG_MK.main.rows_between_operations
            else:
                self.row = self.first_row_of_next_page

    def pushOperationUp(self) -> None:
        """ Сдвиг всех данных операции на строку вверх"""
        self.rowDecrease()
        first_row_of_operation = self.row
        increase = 2
        while first_row_of_operation in self.document_dict:
            first_row_of_operation -= 2
        for row in range(first_row_of_operation, self.row, increase):
            self.document_dict[row] = self.document_dict[row + increase]
        del self.document_dict[self.row]
        if self.isLastRow(first_row_of_operation):
            self.pushOperationUp()
        self.endOfOperation()

    def rowIncrease(self) -> None:
        """ Перейти на следующий ряд """
        self.row += 2

    def rowDecrease(self) -> None:
        """ Перейти на предыдущий ряд """
        self.row -= 2

    @staticmethod
    def page(row: int) -> int:
        """ Определить на какой странице реального
            документа будет находиться строка """
        if row < CFG_MK.row.fp_text_row_first:
            return int((row + CFG_MK.row.row_start_fp) /
                       CFG_MK.row.row_total_fp)
        return int((row + CFG_MK.row.row_start) /
                   CFG_MK.row.row_total)

    def isLastRow(self, row: int) -> bool:
        """ Является ли текущий ряд последним
            на странице реального документа """
        position = row - (self.page(row) - 1) * CFG_MK.row.row_total
        return position == CFG_MK.row.text_row_last

    @property
    def have_space_on_page(self) -> bool:
        """ Возвращает есть ли место для следующей операции на
            текущей странице """
        if self.row < CFG_MK.row.text_row_first:
            crit = CFG_MK.main.free_space_crit_fp
        else:
            crit = CFG_MK.main.free_space_crit
        return self.row_on_cur_page < crit

    @property
    def row_on_cur_page(self) -> int:
        """ Возвращение номера строки при разбиении текста на страницы """
        total_rows = (self.current_page - 1) * CFG_MK.row.row_total
        return int(1 + (self.row - total_rows - CFG_MK.row.row_start) / 2)

    @property
    def first_row_of_next_page(self) -> int:
        """ Возвращает номер строки единого текста, который будет
            соответствовать первой строке следующей строки"""
        return self.current_page * CFG_MK.row.row_total + CFG_MK.row.row_start

    @property
    def current_page(self) -> int:
        """ Возвращает номер страницы активной строки """
        return int((self.row + CFG_MK.row.row_start) / CFG_MK.row.row_total)

    @property
    def document_name(self) -> str:
        """ Наименование документа """
        return self.ws_first_page.cell(row=CFG_MK.row.row_document_name_fp,
                                       column=CFG_MK.col.col_document_name_fp).value

    @document_name.setter
    def document_name(self, value: str) -> None:
        self.ws_template.cell(row=CFG_MK.row.row_document_name,
                              column=CFG_MK.col.col_document_name).value = value
        self.ws_first_page.cell(row=CFG_MK.row.row_document_name_fp,
                                column=CFG_MK.col.col_document_name_fp).value = value

    @property
    def document_deno(self) -> str:
        """ Децимальный номер документа """
        return self.ws_first_page.cell(row=CFG_MK.row.row_document_deno,
                                       column=CFG_MK.col.col_document_deno_fp).value

    @document_deno.setter
    def document_deno(self, value: str) -> None:
        self.ws_template.cell(row=CFG_MK.row.row_document_deno,
                              column=CFG_MK.col.col_document_deno).value = value
        self.ws_first_page.cell(row=CFG_MK.row.row_document_deno,
                                column=CFG_MK.col.col_document_deno_fp).value = value

    @property
    def product_deno(self) -> str:
        """ Децимальный номер изделия """
        return self.ws_first_page.cell(row=CFG_MK.row.row_product_deno,
                                       column=CFG_MK.col.col_product_deno).value

    @product_deno.setter
    def product_deno(self, value: str) -> None:
        self.ws_template.cell(row=CFG_MK.row.row_product_deno,
                              column=CFG_MK.col.col_product_deno).value = value
        self.ws_first_page.cell(row=CFG_MK.row.row_product_deno,
                                column=CFG_MK.col.col_product_deno_fp).value = value

    @property
    def developer(self) -> str:
        """ Разработчик документа """
        return self.ws_first_page.cell(row=CFG_MK.row.row_developer,
                                       column=CFG_MK.col.col_developer).value

    @developer.setter
    def developer(self, value: str) -> None:
        self.ws_first_page.cell(row=CFG_MK.row.row_developer,
                                column=CFG_MK.col.col_developer - 4).value = self.developer
        self.ws_first_page.cell(row=CFG_MK.row.row_developer,
                                column=CFG_MK.col.col_developer).value = value

    @property
    def m_contr(self) -> str:
        """ Метрологический контроль """
        return self.ws_first_page.cell(row=CFG_MK.row.row_m_contr,
                                       column=CFG_MK.col.col_m_contr).value

    @m_contr.setter
    def m_contr(self, value: str) -> None:
        self.ws_first_page.cell(row=CFG_MK.row.row_m_contr,
                                column=CFG_MK.col.col_m_contr - 4).value = self.m_contr
        self.ws_first_page.cell(row=CFG_MK.row.row_m_contr,
                                column=CFG_MK.col.col_m_contr).value = value

    @property
    def checker(self) -> str:
        """ Проверяющий """
        return self.ws_first_page.cell(row=CFG_MK.row.row_checker,
                                       column=CFG_MK.col.col_checker).value

    @checker.setter
    def checker(self, value: str) -> None:
        self.ws_first_page.cell(row=CFG_MK.row.row_checker,
                                column=CFG_MK.col.col_checker - 4).value = self.checker
        self.ws_first_page.cell(row=CFG_MK.row.row_checker,
                                column=CFG_MK.col.col_checker).value = value

    @property
    def approver(self) -> str:
        """ Утверждающий """
        return self.ws_first_page.cell(row=CFG_MK.row.row_approver,
                                       column=CFG_MK.col.col_approver).value

    @approver.setter
    def approver(self, value: str) -> None:
        self.ws_first_page.cell(row=CFG_MK.row.row_approver,
                                column=CFG_MK.col.col_approver - 4).value = self.approver
        self.ws_first_page.cell(row=CFG_MK.row.row_approver,
                                column=CFG_MK.col.col_approver).value = value

    @property
    def n_contr(self) -> str:
        """ Нормоконтроль """
        return self.ws_first_page.cell(row=CFG_MK.row.row_n_contr,
                                       column=CFG_MK.col.col_n_contr).value

    @n_contr.setter
    def n_contr(self, value: str) -> None:
        self.ws_first_page.cell(row=CFG_MK.row.row_n_contr,
                                column=CFG_MK.col.col_n_contr - 4).value = self.n_contr
        self.ws_first_page.cell(row=CFG_MK.row.row_n_contr,
                                column=CFG_MK.col.col_n_contr).value = value

    @property
    def litera(self) -> str:
        """ Литера документа """
        return self.ws_first_page.cell(row=CFG_MK.row.row_litera,
                                       column=CFG_MK.col.col_litera).value

    @litera.setter
    def litera(self, value: str) -> None:
        self.ws_first_page.cell(row=CFG_MK.row.row_litera,
                                column=CFG_MK.col.col_litera).value = value


class EText:
    """ Логическое представление строки
        маршрутной карты"""
    # pylint: disable=too-many-arguments
    # pylint: disable=too-few-public-methods

    def __init__(self, text: str, index: str, col_fst: int, col_end: int | None = None,
                 alignment: str | None = None) -> None:
        self.text = text
        self.index = index
        self.col_fst = col_fst
        self.col_end = col_end
        self.alignment = alignment
        self.merge = bool(self.col_end)
